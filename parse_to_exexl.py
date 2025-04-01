import logging
import os
import pandas as pd
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side
from openpyxl.workbook import Workbook

logging.basicConfig(level=logging.INFO)

def create_empty_excel(columns: list, filename: str, sheet_name: str = 'Sheet1'):
    df = pd.DataFrame(columns=columns)

    if not os.path.exists('bot/handlers/excel_files'):
        os.makedirs('bot/handlers/excel_files')

    filepath = os.path.join('bot/handlers/excel_files', filename)
    excel_writer = pd.ExcelWriter(filepath, engine='xlsxwriter')
    df.to_excel(excel_writer, index=False, sheet_name=sheet_name, freeze_panes=(1, 0))
    excel_writer._save()

    return filepath


def create_tabel_users():
    filepath = create_empty_excel(
        columns=['Регистрационный номер', 'Код ОКПО', 'ИНН', 'Область/город республиканского значения',
                 'Район/город областного значения', 'Микрорайон/жилмассив',
                 'Улица (проспект, бульвар, переулок и т.п.)', '№ дома', '№ квартиры (офиса, комнаты и т.п.)',
                 'Телефон', 'Факс', 'Электронный адрес (e-mail)', 'Дата приказа', '', 'Дата первичной регистрации',
                 'Форма собственности', 'Фамилия, имя, отчество', 'Основной вид деятельности',
                 'Код экономической деятельности', 'лиц', 'Общее количество учредителей (участников)',
                 'Учредители (участники):', '(участник) Ф.И.О.'],
        filename='users.xlsx')


def parse_excel_to_dict_list(filepath: str, sheet_name='Sheet1'):
    # Загружаем Excel файл в DataFrame
    df = pd.read_excel(filepath, sheet_name=sheet_name)

    # Преобразуем DataFrame в список словарей
    dict_list = df.to_dict(orient='records')

    return dict_list


def create_excel_from_dict_list(dict_list: list, output_filename: str, sheet_name='Sheet1'):
    if not os.path.exists('bot/handlers/excel_files'):
        os.makedirs('bot/handlers/excel_files')
    filepath = os.path.join('bot/handlers/excel_files', output_filename)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Записываем данные из списка словарей в Excel
    if dict_list:
        header = list(dict_list[0].keys())
        ws.append(header)  # Записываем заголовки

        for row in dict_list:
            ws.append([row[col] for col in header])

    # Настраиваем стили для красивого вида
    header_style = NamedStyle(name='header')
    header_style.font = Font(bold=True, color='FFFFFF')
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    header_style.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    border_style = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    header_style.border = border_style

    cell_style = NamedStyle(name='cell')
    cell_style.alignment = Alignment(horizontal='left', vertical='center')
    cell_style.border = border_style

    for cell in ws[1]:  # Применяем стиль к заголовкам
        cell.style = header_style

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.style = cell_style

    # Автоматическое изменение ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    # Сохраняем файл
    wb.save(filepath)
    logging.info("Файл сохранён")
    return filepath

